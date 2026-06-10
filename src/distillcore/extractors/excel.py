"""Excel extraction via openpyxl. Requires distillcore[excel]."""

from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl

from ..models import ExtractionResult, PageText


class ExcelExtractor:
    """Extract tabular text from .xlsx workbooks.

    Each non-empty worksheet becomes one ``PageText`` (page_number = 1-based
    sheet index). Cells within a row are tab-separated; rows are newline-
    separated. Cell values are coerced to strings (datetimes via isoformat,
    everything else via ``str``); type inference is deliberately out of scope
    for v1 — let the downstream pipeline classify the data.

    Output format example::

        col_a\\tcol_b\\tcol_c
        1\\t2\\t3
        4\\t5\\t6

    Metadata populated on ``ExtractionResult.metadata``:
        - ``sheet_names`` — list of all sheet titles in workbook order
        - ``row_counts`` — dict of sheet name → number of non-empty rows
          (including header)

    Limitations:
        - ``.xls`` (legacy binary) is not supported. Use ``.xlsx`` instead.
        - Formulas are read as cached values (``data_only=True``). If the
          workbook hasn't been opened in Excel since the formulas were
          written, cached values may be ``None``.
        - Empty sheets are skipped entirely (no PageText emitted).
    """

    formats = ["xlsx"]

    def extract(self, source: Path | str, config: Any = None) -> ExtractionResult:
        wb = openpyxl.load_workbook(str(source), read_only=True, data_only=True)
        try:
            sheet_names = list(wb.sheetnames)
            pages: list[PageText] = []
            row_counts: dict[str, int] = {}
            page_number = 1
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                rows = list(_iter_non_empty_rows(ws))
                row_counts[sheet_name] = len(rows)
                if not rows:
                    continue
                text = "\n".join("\t".join(row) for row in rows)
                pages.append(PageText(page_number=page_number, text=text))
                page_number += 1
        finally:
            wb.close()

        full_text = "\n\n".join(p.text for p in pages)
        return ExtractionResult(
            pages=pages,
            full_text=full_text,
            page_count=len(pages),
            format="xlsx",
            metadata={
                "sheet_names": sheet_names,
                "row_counts": row_counts,
            },
        )


def _cell_to_str(value: Any) -> str:
    """Coerce an Excel cell value to a plain string."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, bool):
        # bool must be checked before int (it's a subclass).
        return "TRUE" if value else "FALSE"
    return str(value)


def _iter_non_empty_rows(ws: Any) -> Any:
    """Yield each row's cell values as a list[str], skipping fully-empty rows."""
    for row in ws.iter_rows(values_only=True):
        cells = [_cell_to_str(v) for v in row]
        # Skip rows where every cell is empty after coercion.
        if any(c for c in cells):
            yield cells
